"""Build a formatted Excel workbook from the structured JSON returned by Gemini."""

from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
_THIN_SIDE = Side(style="thin", color="000000")
_TABLE_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Locale-aware fixed labels
# ---------------------------------------------------------------------------
_LABELS: dict[str, dict[str, str]] = {
    "zh": {
        "address": "地址",
        "phone": "电话",
        "fax": "传真",
        "counterparty": "客户/供应商",
        "doc_number": "单号",
        "date": "日期",
        "month": "月份",
        "total": "合计",
        "tax_rate": "税率",
        "tax": "税金",
        "grand_total": "总计（含税）",
    },
    "en": {
        "address": "Address",
        "phone": "Phone",
        "fax": "Fax",
        "counterparty": "Customer/Supplier",
        "doc_number": "Quote No.",
        "date": "Date",
        "month": "Month",
        "total": "Total",
        "tax_rate": "Tax Rate",
        "tax": "Tax",
        "grand_total": "Total (incl. tax)",
    },
}


def _normalize_locale(locale: str | None) -> str:
    """Map any locale string to a supported label set (default zh)."""
    if locale and locale.strip().lower().startswith("en"):
        return "en"
    return "zh"


def _try_number(value: Any) -> int | float | str:
    """Try to convert *value* to a number, else return it as string."""
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return ""
        try:
            return int(s)
        except ValueError:
            pass
        try:
            return float(s)
        except ValueError:
            pass
    return str(value) if value is not None else ""


def _col_letter(col_index: int) -> str:
    """1-based col index → letter."""
    return get_column_letter(col_index)


def build_excel(data: dict[str, Any], locale: str | None = None) -> bytes:
    """Return an in-memory xlsx file as *bytes* from the structured JSON.

    ``locale`` selects the language of the fixed labels (meta/summary
    headers). Supported: "zh" (default) and "en".
    """

    loc = _normalize_locale(locale)
    labels = _LABELS[loc]

    wb = Workbook()
    ws = wb.active
    ws.title = (data.get("title") or "Sheet1").strip()[:31]  # Excel 31-char limit

    headers: list[str] = (data.get("table") or {}).get("headers", [])
    num_cols = max(len(headers), 1)

    current_row = 1

    # ------------------------------------------------------------------
    # 1. Company info area
    # ------------------------------------------------------------------
    company: dict[str, Any] = data.get("companyInfo") or {}
    company_name = (company.get("name") or "").strip()
    if company_name:
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=num_cols,
        )
        cell = ws.cell(row=current_row, column=1, value=company_name)
        cell.font = Font(size=16, bold=True)
        cell.alignment = _CENTER
        current_row += 1

    # secondary company lines (address, phone, fax, other)
    secondary_lines: list[str] = []
    for key in ("address", "phone", "fax"):
        val = (company.get(key) or "").strip()
        if val:
            secondary_lines.append(f"{labels.get(key, key)}: {val}")
    for line in company.get("other") or []:
        if isinstance(line, str) and line.strip():
            secondary_lines.append(line.strip())

    if secondary_lines:
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=num_cols,
        )
        cell = ws.cell(row=current_row, column=1, value="  ".join(secondary_lines))
        cell.font = Font(size=9, color="666666")
        cell.alignment = _CENTER
        current_row += 1

    # ------------------------------------------------------------------
    # 2. Document title
    # ------------------------------------------------------------------
    title = (data.get("title") or "").strip()
    if title:
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=num_cols,
        )
        cell = ws.cell(row=current_row, column=1, value=title)
        cell.font = Font(size=14, bold=True)
        cell.alignment = _CENTER
        current_row += 1
        current_row += 1  # blank row after title

    # ------------------------------------------------------------------
    # 3. Meta info (document number, date, counterparty …)
    # ------------------------------------------------------------------
    meta: dict[str, Any] = data.get("documentMeta") or {}
    counterparty: dict[str, Any] = data.get("counterparty") or {}
    meta_parts: list[str] = []

    cp_name = (counterparty.get("name") or "").strip()
    if cp_name:
        meta_parts.append(f"{labels['counterparty']}: {cp_name}")
    doc_num = (meta.get("documentNumber") or "").strip()
    if doc_num:
        meta_parts.append(f"{labels['doc_number']}: {doc_num}")
    doc_date = (meta.get("date") or "").strip()
    if doc_date:
        meta_parts.append(f"{labels['date']}: {doc_date}")
    doc_month = (meta.get("month") or "").strip()
    if doc_month:
        meta_parts.append(f"{labels['month']}: {doc_month}")
    for k, v in (meta.get("otherMeta") or {}).items():
        if isinstance(v, str) and v.strip():
            meta_parts.append(f"{k}: {v}")

    if meta_parts:
        # distribute meta across cells evenly
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=num_cols,
        )
        cell = ws.cell(row=current_row, column=1, value="    ".join(meta_parts))
        cell.font = Font(size=10)
        cell.alignment = _LEFT
        current_row += 1
        current_row += 1  # blank row

    # ------------------------------------------------------------------
    # 4. Table header
    # ------------------------------------------------------------------
    table_data: dict[str, Any] = data.get("table") or {}
    rows: list[list[Any]] = table_data.get("rows", [])
    formula_cols: list[dict[str, Any]] = table_data.get("formulaColumns") or []

    # Build a lookup: col_index → formula template
    formula_map: dict[int, str] = {}
    for fc in formula_cols:
        idx = fc.get("colIndex")
        formula = fc.get("formula", "")
        if idx is not None and formula:
            formula_map[int(idx)] = formula

    header_row = current_row
    if headers:
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=ci, value=h)
            cell.font = Font(bold=True, size=10)
            cell.fill = _HEADER_FILL
            cell.border = _TABLE_BORDER
            cell.alignment = _CENTER
        current_row += 1

    # ------------------------------------------------------------------
    # 5. Table data rows
    # ------------------------------------------------------------------
    data_start_row = current_row
    for row_values in rows:
        if not isinstance(row_values, list):
            continue
        for ci in range(1, num_cols + 1):
            raw = row_values[ci - 1] if ci - 1 < len(row_values) else ""

            # Check if this column should have a formula
            # colIndex in the JSON is 0-based
            if (ci - 1) in formula_map:
                tmpl = formula_map[ci - 1]
                # Replace {row} with the actual Excel row number
                formula_str = tmpl.replace("{row}", str(current_row))
                # Ensure it starts with '='
                if formula_str.startswith("="):
                    cell = ws.cell(row=current_row, column=ci)
                    cell.value = formula_str
                else:
                    cell = ws.cell(row=current_row, column=ci, value=_try_number(raw))
            else:
                cell = ws.cell(row=current_row, column=ci, value=_try_number(raw))

            cell.border = _TABLE_BORDER
            cell.alignment = _CENTER
            cell.font = Font(size=10)
        current_row += 1

    # ------------------------------------------------------------------
    # 6. Summary rows
    # ------------------------------------------------------------------
    summary: dict[str, Any] = data.get("summary") or {}
    summary_lines: list[tuple[str, Any]] = []

    total_label = (summary.get("totalLabel") or "").strip()
    if loc == "en" and total_label in ("合计", "总计", "合计/总计"):
        # AI may echo the Chinese label from the source document; keep the
        # EN output consistent with the selected locale.
        total_label = labels["total"]
    total_amount = summary.get("totalAmount")
    if total_label or total_amount is not None:
        summary_lines.append((total_label or labels["total"], total_amount if total_amount is not None else ""))

    tax_rate = summary.get("taxRate")
    tax_amount = summary.get("taxAmount")
    if tax_rate or tax_amount is not None:
        label = f"{labels['tax_rate']} {tax_rate}" if tax_rate else labels["tax"]
        summary_lines.append((label, tax_amount if tax_amount is not None else ""))

    grand_total = summary.get("grandTotal")
    if grand_total is not None:
        summary_lines.append((labels["grand_total"], grand_total))

    for k, v in (summary.get("otherSummary") or {}).items():
        summary_lines.append((str(k), v))

    for label, value in summary_lines:
        # merge label area, put value in last column
        if num_cols >= 2:
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=num_cols - 1,
            )
            cell = ws.cell(row=current_row, column=1, value=label)
            cell.font = Font(bold=True, size=10)
            cell.alignment = _RIGHT
            cell.border = _TABLE_BORDER

            val_cell = ws.cell(
                row=current_row, column=num_cols, value=_try_number(value)
            )
            val_cell.font = Font(bold=True, size=10)
            val_cell.alignment = _CENTER
            val_cell.border = _TABLE_BORDER
        else:
            cell = ws.cell(row=current_row, column=1, value=f"{label}: {value}")
            cell.font = Font(bold=True, size=10)
            cell.border = _TABLE_BORDER
        current_row += 1

    current_row += 1  # blank row

    # ------------------------------------------------------------------
    # 7. Footer (notes + signatures)
    # ------------------------------------------------------------------
    footer: dict[str, Any] = data.get("footer") or {}
    notes: list[str] = footer.get("notes") or []
    signatures: list[str] = footer.get("signatures") or []

    for note in notes:
        if isinstance(note, str) and note.strip():
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=num_cols,
            )
            cell = ws.cell(row=current_row, column=1, value=note.strip())
            cell.font = Font(size=9, color="666666")
            cell.alignment = _LEFT
            current_row += 1

    if signatures:
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=num_cols,
        )
        cell = ws.cell(row=current_row, column=1, value="    ".join(signatures))
        cell.font = Font(size=9)
        cell.alignment = _LEFT
        current_row += 1

    # ------------------------------------------------------------------
    # 8. Auto column widths
    # ------------------------------------------------------------------
    for ci in range(1, num_cols + 1):
        max_len = 8  # minimum width
        for row_cells in ws.iter_rows(min_col=ci, max_col=ci, min_row=1, max_row=current_row):
            for c in row_cells:
                if c.value is not None:
                    # Estimate width: CJK chars count as ~2
                    text = str(c.value)
                    width = 0
                    for ch in text:
                        width += 2 if ord(ch) > 127 else 1
                    max_len = max(max_len, min(width + 2, 40))
        ws.column_dimensions[_col_letter(ci)].width = max_len

    # ------------------------------------------------------------------
    # Write to bytes
    # ------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
