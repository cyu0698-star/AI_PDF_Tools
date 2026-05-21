"""Serialize an Excel template for AI prompting & fill it with AI-returned mappings."""

from __future__ import annotations

import io
import re
from copy import copy
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, coordinate_to_tuple


def serialize_template(xlsx_bytes: bytes) -> str:
    """Read an xlsx template and produce a compact text description for the AI prompt.

    Annotates each cell with its role:
      - [LABEL]       fixed text ending with ：/: — never overwrite
      - [LABEL+VALUE] merged cell with label pattern + space for values — fill values after labels
      - [PLACEHOLDER] placeholder text like 'xxx' — overwrite entirely
      - [EMPTY]       empty cell — fillable
      - [DATA]        sequence number or other data
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    lines: list[str] = []

    lines.append(f"Sheet: {ws.title}")
    lines.append(f"Dimensions: {ws.dimensions}")

    # Merged regions
    merged_ranges = list(ws.merged_cells.ranges)
    merge_map: dict[tuple[int, int], str] = {}
    if merged_ranges:
        lines.append("Merged regions:")
        for mr in merged_ranges:
            lines.append(f"  {mr}")
            merge_map[(mr.min_row, mr.min_col)] = str(mr)

    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    # Walk rows, collapsing consecutive empty rows
    empty_run_start: int | None = None

    for row_idx in range(1, max_row + 1):
        row_parts: list[str] = []
        row_has_content = False

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue

            coord = cell.coordinate
            val = cell.value
            is_merged = (row_idx, col_idx) in merge_map

            if val is None or (isinstance(val, str) and not val.strip()):
                row_parts.append(f"{coord} [EMPTY]")
                continue

            row_has_content = True
            val_str = str(val).strip()
            tag = _classify_cell(val_str, is_merged)
            merge_note = f" (merged {merge_map[(row_idx, col_idx)]})" if is_merged else ""
            row_parts.append(f"{coord}={val!r} {tag}{merge_note}")

        if not row_has_content:
            if empty_run_start is None:
                empty_run_start = row_idx
            continue

        # Flush empty run
        if empty_run_start is not None:
            end = row_idx - 1
            if end == empty_run_start:
                lines.append(f"Row {empty_run_start}: (all empty, fillable)")
            else:
                lines.append(f"Rows {empty_run_start}-{end}: (all empty, fillable)")
            empty_run_start = None

        lines.append(f"Row {row_idx}: {', '.join(row_parts)}")

    # Trailing empty rows
    if empty_run_start is not None:
        end = max_row
        if end == empty_run_start:
            lines.append(f"Row {empty_run_start}: (all empty, fillable)")
        else:
            lines.append(f"Rows {empty_run_start}-{end}: (all empty, fillable)")

    wb.close()
    return "\n".join(lines)


def _classify_cell(val_str: str, is_merged: bool) -> str:
    """Classify a cell value into a role tag."""
    # Placeholder like 'xxx有限公司'
    if re.search(r"xxx|XXX|占位|placeholder", val_str, re.IGNORECASE):
        return "[PLACEHOLDER]"

    # Merged cell with label:value pattern — treat as LABEL+VALUE
    if is_merged and re.search(r"[：:]", val_str):
        return "[LABEL+VALUE]"

    # Non-merged label: text ending with ： or :
    stripped = val_str.rstrip()
    if re.match(r"^.{1,10}[：:]\s*$", stripped):
        return "[LABEL]"

    # Pure number (sequence numbers, amounts)
    try:
        float(val_str)
        return "[DATA]"
    except ValueError:
        pass

    return "[FIXED]"


# ---------------------------------------------------------------------------
# fill_template — with auto row expansion
# ---------------------------------------------------------------------------

def fill_template(xlsx_bytes: bytes, mapping: dict[str, Any]) -> bytes:
    """Open the original template and fill cells according to the AI mapping.

    When tableRegion.rows exceeds the template's data slots (endDataRow − startRow + 1),
    new rows are inserted with formatting copied from the last data row, and
    summaryCells refs are shifted down accordingly.
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    def _write_cell(ref: str, value: Any) -> None:
        cell = ws[ref]
        if isinstance(cell, MergedCell):
            for mr in ws.merged_cells.ranges:
                if ref in mr:
                    cell = ws.cell(row=mr.min_row, column=mr.min_col)
                    break
        cell.value = _coerce(value)

    # 1. Single cells
    for entry in mapping.get("cells") or []:
        ref = entry.get("ref")
        val = entry.get("value")
        if ref:
            _write_cell(ref, val)

    # 2. Table region (batch rows) — with auto-expansion
    table = mapping.get("tableRegion") or {}
    start_ref = table.get("startRef")
    table_rows: list[list[Any]] = table.get("rows") or []
    rows_inserted = 0

    if start_ref and table_rows:
        start_row, start_col = coordinate_to_tuple(start_ref)

        # Auto-detect endDataRow from template: scan the sequence-number
        # column (col before startRef) downward from start_row, find the
        # last row that has a numeric value (1,2,3...).
        end_data_row = _detect_end_data_row(ws, start_row, start_col)

        if end_data_row and end_data_row >= start_row:
            template_slots = end_data_row - start_row + 1
        else:
            template_slots = len(table_rows)  # cannot detect, skip expansion

        extra_rows = len(table_rows) - template_slots
        if extra_rows > 0 and end_data_row:
            insert_at = end_data_row + 1
            _insert_rows_with_style(ws, insert_at, extra_rows, end_data_row)
            rows_inserted = extra_rows

            # Extend sequence numbers in column(s) before startRef
            for ci in range(1, start_col):
                last_val = ws.cell(row=end_data_row, column=ci).value
                if isinstance(last_val, (int, float)):
                    for i in range(extra_rows):
                        ws.cell(
                            row=end_data_row + 1 + i, column=ci,
                            value=int(last_val) + 1 + i,
                        )

        # Write data into cells
        for ri, row_data in enumerate(table_rows):
            for ci, val in enumerate(row_data):
                ws.cell(
                    row=start_row + ri, column=start_col + ci,
                    value=_coerce(val),
                )

    # 3. Summary cells — shift refs down by inserted rows
    for entry in mapping.get("summaryCells") or []:
        ref = entry.get("ref")
        val = entry.get("value")
        if ref:
            if rows_inserted > 0:
                ref = _shift_ref(ref, rows_inserted)
            _write_cell(ref, val)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _detect_end_data_row(ws: Any, start_row: int, start_col: int) -> int | None:
    """Auto-detect the last data row in the template.

    Strategy: look at the column(s) *before* startRef (typically the sequence
    number column A). Walk down from start_row; the last row with a numeric
    value (1, 2, 3, …) is the end of the data region.

    Falls back to scanning start_col itself if no sequence column found.
    """
    # Try sequence-number column (columns before start_col)
    for seq_col in range(1, start_col):
        last_data_row = None
        for row_idx in range(start_row, (ws.max_row or start_row) + 1):
            val = ws.cell(row=row_idx, column=seq_col).value
            if isinstance(val, (int, float)):
                last_data_row = row_idx
            else:
                # Stop at first non-numeric (skipping None for safety)
                if val is not None and str(val).strip():
                    break
                # Also stop if the *data* columns are non-empty with labels
                data_val = ws.cell(row=row_idx, column=start_col).value
                if data_val is not None and str(data_val).strip():
                    break
        if last_data_row is not None:
            return last_data_row

    # Fallback: scan startRef column for empty-row boundary
    for row_idx in range(start_row, (ws.max_row or start_row) + 1):
        val = ws.cell(row=row_idx, column=start_col).value
        if val is not None and str(val).strip():
            continue
        # Check if the *whole* row from start_col is empty (template slot)
        all_empty = True
        for ci in range(start_col, (ws.max_column or start_col) + 1):
            cv = ws.cell(row=row_idx, column=ci).value
            if cv is not None and str(cv).strip():
                all_empty = False
                break
        if not all_empty:
            return row_idx - 1 if row_idx > start_row else None
    return None


def _insert_rows_with_style(ws: Any, insert_at: int, count: int, style_source_row: int) -> None:
    """Insert *count* rows at *insert_at*, copying cell styles from *style_source_row*."""
    ws.insert_rows(insert_at, count)
    max_col = ws.max_column or 1
    for offset in range(count):
        target_row = insert_at + offset
        for col_idx in range(1, max_col + 1):
            src_cell = ws.cell(row=style_source_row, column=col_idx)
            tgt_cell = ws.cell(row=target_row, column=col_idx)
            if src_cell.has_style:
                tgt_cell.font = copy(src_cell.font)
                tgt_cell.border = copy(src_cell.border)
                tgt_cell.fill = copy(src_cell.fill)
                tgt_cell.number_format = src_cell.number_format
                tgt_cell.alignment = copy(src_cell.alignment)
                tgt_cell.protection = copy(src_cell.protection)
        if ws.row_dimensions[style_source_row].height:
            ws.row_dimensions[target_row].height = ws.row_dimensions[style_source_row].height


def _shift_ref(ref: str, delta: int) -> str:
    """Shift a cell reference down by *delta* rows.  E.g. 'F19' + 3 → 'F22'."""
    m = re.match(r"^([A-Z]+)(\d+)$", ref)
    if not m:
        return ref
    return f"{m.group(1)}{int(m.group(2)) + delta}"


def _coerce(value: Any) -> int | float | str | None:
    """Try to convert value to a number, otherwise return as string."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return int(s)
        except ValueError:
            pass
        try:
            return float(s)
        except ValueError:
            pass
        return s
    return str(value)
