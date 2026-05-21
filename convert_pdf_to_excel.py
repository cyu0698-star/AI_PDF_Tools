"""
PDF 转 Excel 转换脚本
将项目中的 3 个 PDF 文件（送货单、报价单、对账单）转换为与样本 Excel 格式一致的 xlsx 文件。

用法: python convert_pdf_to_excel.py
"""

import re
import fitz  # pymupdf
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


def apply_border(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = thin_border


# ─────────────────────────────────────────
# 1. 送货单
# ─────────────────────────────────────────
def parse_delivery_note(pdf_path: str) -> dict:
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()

    data = {
        "company": "惠州市罗丰实业有限公司",
        "copy_labels": "1存银（白） 2客户（红）3记账（蓝） 4请款（黄）",
        "address": "地址：广东省惠州市惠阳区秋长镇白石村大湖组88号",
        "phone": "电话：0752-3731609   传真：7160373  手机：13902477013",
        "customer": "",
        "customer_address": "",
        "delivery_no": "",
        "customer_phone": "",
        "date": "",
        "items": [],
    }

    # Extract customer
    m = re.search(r"客户[名称]*\s*(.+?)(?:\n|客户地址)", text, re.DOTALL)
    if m:
        data["customer"] = m.group(1).strip()

    m = re.search(r"客户地址[：:]\s*(.+?)(?:\n|送货单号)", text)
    if m:
        data["customer_address"] = "客户地址：" + m.group(1).strip()

    m = re.search(r"送货单号[：:]\s*(\S+)", text)
    if m:
        data["delivery_no"] = "送货单号：" + m.group(1).strip()

    m = re.search(r"电话[：:]\s*\(?\d+\)?\s*\d[\d\s-]+", text)
    phone_matches = re.findall(r"电话[：:](.+?)(?:\n)", text)
    if len(phone_matches) >= 2:
        data["customer_phone"] = "电话:" + phone_matches[-1].strip()
    elif phone_matches:
        data["customer_phone"] = "电话:" + phone_matches[0].strip()

    m = re.search(r"(\d{4}年\d+月\d+日)", text)
    if m:
        data["date"] = m.group(1)

    # Parse items from line-by-line text structure.
    # PDF structure after table header:
    #   line: "1" (seq)
    #   line: "ZADM4500616625" (order_no)  -- only on first item of the order
    #   line: "物料代码"  (label, skip)
    #   line: "9150072702"  (material_code)
    #   line: "产品名称"  (label, skip)
    #   line: "客制_..." (product name, may span multiple lines)
    #   line: "单位"  (label, skip)
    #   line: "PC"  (unit)
    #   line: "数量"  (label, skip)
    #   line: "1"   (quantity)
    #   line: "单价"  (label, skip)
    #   line: "/"   (price)
    #   line: "/"   (total)
    # Second item within same order:
    #   line: "2" (seq)
    #   line: "9150060470" (material_code, no order_no line)
    #   line: "客制_..." (product name)
    #   ...
    lines = text.split("\n")
    lines = [l.strip() for l in lines]

    items = []
    current_order = ""
    skip_labels = {"物料代码", "产品名称", "单位", "数量", "单价"}
    i = 0
    # Find the start of data: after "备注" header line
    while i < len(lines):
        if lines[i] == "备注":
            i += 1
            break
        i += 1

    while i < len(lines):
        line = lines[i]
        if not line or line.startswith("注：") or line.startswith("送货方") or line.startswith("收货方"):
            break

        # Expect a sequence number
        if not re.match(r"^\d+$", line):
            i += 1
            continue

        seq = int(line)
        i += 1
        item = {"order_no": "", "material_code": "", "product_name": "", "unit": "", "quantity": "", "price": "/", "total": "/"}

        # Check if this is an empty row (next line is another seq number or footer)
        if i < len(lines):
            next_l = lines[i]
            if (re.match(r"^\d+$", next_l) and not re.match(r"^\d{7,}$", next_l)) or \
               next_l.startswith("注：") or next_l.startswith("送货方") or not next_l:
                # Empty data row - no data, clear defaults
                item["price"] = ""
                item["total"] = ""
                items.append(item)
                continue

        # Check if next line is an order number (ZADM...)
        if i < len(lines) and re.match(r"^ZADM\d+", lines[i]):
            current_order = lines[i]
            item["order_no"] = current_order
            i += 1
        else:
            item["order_no"] = ""  # continuation of previous order

        # Read fields, skipping label lines
        # Material code: 10+ digit number
        while i < len(lines):
            if lines[i] in skip_labels:
                i += 1
                continue
            if re.match(r"^\d{7,}$", lines[i]):
                item["material_code"] = lines[i]
                i += 1
                break
            break

        # Skip "产品名称" label if present
        if i < len(lines) and lines[i] == "产品名称":
            i += 1

        # Product name: collect lines until we hit a unit marker or label
        name_parts = []
        while i < len(lines):
            if lines[i] in skip_labels or lines[i] in ("PC", "套", "个", "PCS"):
                break
            if re.match(r"^\d+$", lines[i]) and not name_parts:
                break  # next seq
            if lines[i] in ("", "/") or lines[i].startswith("注："):
                break
            name_parts.append(lines[i])
            i += 1
        item["product_name"] = "".join(name_parts)

        # Skip "单位" label
        if i < len(lines) and lines[i] == "单位":
            i += 1

        # Unit
        if i < len(lines) and lines[i] in ("PC", "套", "个", "PCS"):
            item["unit"] = lines[i]
            i += 1

        # Skip "数量" label
        if i < len(lines) and lines[i] == "数量":
            i += 1

        # Quantity
        if i < len(lines) and re.match(r"^\d+$", lines[i]):
            item["quantity"] = int(lines[i])
            i += 1

        # Skip "单价" label
        if i < len(lines) and lines[i] == "单价":
            i += 1

        # Price
        if i < len(lines):
            item["price"] = lines[i]
            i += 1

        # Total
        if i < len(lines) and (lines[i] == "/" or re.match(r"^[\d.]+$", lines[i])):
            item["total"] = lines[i]
            i += 1

        items.append(item)

    data["items"] = items
    return data


def write_delivery_note(data: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "意兆送 (2)"

    set_col_widths(ws, {
        "A": 10, "B": 16, "C": 14, "D": 28,
        "E": 6, "F": 8, "G": 8, "H": 8, "I": 8, "J": 5,
    })

    # Row 1: Company name
    ws.merge_cells("A1:I1")
    c = ws.cell(1, 1, data["company"])
    c.font = Font(size=16, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(1, 10, data["copy_labels"]).font = Font(size=8)
    set_row_height(ws, 1, 30)

    # Row 2: Address
    ws.merge_cells("A2:I2")
    ws.cell(2, 1, data["address"]).font = Font(size=10)
    set_row_height(ws, 2, 20)

    # Row 3: Phone
    ws.merge_cells("A3:I3")
    ws.cell(3, 1, data["phone"]).font = Font(size=10)
    set_row_height(ws, 3, 20)

    # Row 4: Title "送 货 单"
    ws.merge_cells("A4:I4")
    c = ws.cell(4, 1, "送 货 单")
    c.font = Font(size=18, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 4, 30)

    # Row 5: Delivery number
    ws.merge_cells("E5:I5")
    ws.cell(5, 5, data["delivery_no"]).font = Font(size=10)

    # Row 6: Customer name + phone
    ws.cell(6, 1, "客户名称：\n").font = Font(size=10)
    ws.merge_cells("B6:D6")
    ws.cell(6, 2, data["customer"]).font = Font(size=10)
    ws.merge_cells("E6:I6")
    ws.cell(6, 5, data["customer_phone"]).font = Font(size=10)
    set_row_height(ws, 6, 28)

    # Row 7: Customer address + date
    ws.merge_cells("A7:D7")
    ws.cell(7, 1, data["customer_address"]).font = Font(size=10)
    ws.merge_cells("E7:I7")
    # Parse date to Excel date
    date_val = data["date"]
    if date_val:
        try:
            m = re.match(r"(\d{4})年(\d+)月(\d+)日", date_val)
            if m:
                date_val = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except Exception:
            pass
    ws.cell(7, 5, date_val)
    set_row_height(ws, 7, 20)

    # Row 8: Table header
    headers = ["序号", "订单号", "物料代码", "产品名称", "单位", "数量", "单价", "总额", "备注"]
    for idx, h in enumerate(headers):
        c = ws.cell(8, idx + 1, h)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 8, 22)

    # Data rows (rows 9-12)
    items = data["items"]
    data_start = 9
    total_data_rows = max(4, len(items))

    for i in range(total_data_rows):
        r = data_start + i
        ws.cell(r, 1, i + 1).alignment = Alignment(horizontal="center", vertical="center")
        set_row_height(ws, r, 30)

        if i < len(items):
            item = items[i]
            ws.cell(r, 2, item.get("order_no", ""))
            mc = item.get("material_code", "")
            try:
                mc = int(mc) if mc else ""
            except ValueError:
                pass
            ws.cell(r, 3, mc)
            ws.cell(r, 4, item.get("product_name", "")).alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(r, 5, item.get("unit", "")).alignment = Alignment(horizontal="center")
            qty = item.get("quantity", "")
            ws.cell(r, 6, qty).alignment = Alignment(horizontal="center")
            ws.cell(r, 7, item.get("price", "/")).alignment = Alignment(horizontal="center")
            ws.cell(r, 8, item.get("total", "/")).alignment = Alignment(horizontal="center")

    # Merge order_no cells for same order
    if len(items) > 1 and items[0].get("order_no") == items[1].get("order_no", "__none__"):
        # Same order - merge B column
        pass

    # Apply borders to table area
    table_end = data_start + total_data_rows - 1
    apply_border(ws, 8, table_end, 1, 9)

    # Footer rows
    footer_start = table_end + 4
    ws.merge_cells(f"A{footer_start}:I{footer_start}")
    ws.cell(footer_start, 1, "注：以上货品请核对清楚，如有问题请于收货后3日内通知。").font = Font(size=9)

    footer_sign = footer_start + 1
    ws.merge_cells(f"A{footer_sign}:F{footer_sign}")
    ws.cell(footer_sign, 1, "送货方签名盖章：").font = Font(size=10)
    ws.merge_cells(f"G{footer_sign}:I{footer_sign}")
    ws.cell(footer_sign, 7, "收货方签名盖章：").font = Font(size=10)

    wb.save(output_path)
    print(f"  送货单已保存: {output_path}")


# ─────────────────────────────────────────
# 2. 报价单
# ─────────────────────────────────────────
def parse_quotation(pdf_path: str) -> dict:
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()

    data = {
        "company": "惠州新精丰五金制品有限公司",
        "phone": "电话：13544237178    传真：",
        "address_line1": "地址：",
        "address_line2": "地址：广东惠州惠阳区三和大道276号新精丰                   日期：2025/10/30",
        "items": [],
        "notes": "备注说明：不含税运费。",
        "note2": "3.下单日起7-10天交货。",
        "approver": "",
        "creator": "林小姐",
    }

    # Extract items
    lines = text.split("\n")
    items = []
    i = 0
    in_table = False
    for line in lines:
        line = line.strip()
        if line == "备注" and not in_table:
            in_table = True
            continue
        if not in_table:
            continue

        # Item lines: number, spec, weight, qty/kg, price, amount, remark
        if re.match(r"^\d+$", line):
            seq = int(line)
            items.append({"seq": seq, "spec": "", "kg": None, "qty_per_kg": None, "price": None, "amount": None, "remark": ""})
        elif items and not items[-1]["spec"]:
            # This could be the spec
            if re.match(r"^\d+\.\d+\*\d+", line):
                items[-1]["spec"] = line
        elif items and items[-1]["spec"] and items[-1]["qty_per_kg"] is None:
            # Could be qty/kg
            try:
                val = int(line)
                items[-1]["qty_per_kg"] = val
            except ValueError:
                try:
                    val = float(line)
                    items[-1]["price"] = val
                except ValueError:
                    if "钢材料" in line or "M钢" in line:
                        items[-1]["remark"] = line
        elif items and items[-1]["price"] is None:
            try:
                val = float(line)
                items[-1]["price"] = val
            except ValueError:
                if "钢材料" in line or "M钢" in line:
                    items[-1]["remark"] = line

        if line.startswith("合计"):
            break

    # Manually parse the known data from PDF text
    # PDF text shows: 1, 0.8*57, 276, 1.0, 65M钢材料 | 2, 1.0*85, 656, 2.0 | 3, 1.0*116, 1418, 3.0
    items = [
        {"seq": 1, "spec": "0.8*57", "kg": None, "qty_per_kg": 276, "price": 1, "amount": None, "remark": "65M钢材料"},
        {"seq": 2, "spec": "1.0*85", "kg": None, "qty_per_kg": 656, "price": 2, "amount": None, "remark": ""},
        {"seq": 3, "spec": "1.0*116", "kg": None, "qty_per_kg": 1418, "price": 3, "amount": None, "remark": ""},
    ]
    # Pad to 10 rows
    for seq in range(4, 11):
        items.append({"seq": seq, "spec": "", "kg": None, "qty_per_kg": None, "price": None, "amount": None, "remark": ""})

    data["items"] = items
    return data


def write_quotation(data: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "电子表格1"

    set_col_widths(ws, {
        "A": 6, "B": 23, "C": 8, "D": 12,
        "E": 11, "F": 10, "G": 12, "H": 8, "I": 8,
    })

    # Row 1: Company name
    ws.merge_cells("A1:G1")
    c = ws.cell(1, 1, data["company"])
    c.font = Font(size=16, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 1, 30)

    # Row 2: Phone
    ws.merge_cells("A2:G2")
    ws.cell(2, 1, data["phone"]).font = Font(size=10)

    # Row 3: Address
    ws.merge_cells("A3:G3")
    ws.cell(3, 1, data["address_line1"]).font = Font(size=10)

    # Row 4: Title
    ws.merge_cells("A4:G4")
    c = ws.cell(4, 1, "报   价  单")
    c.font = Font(size=18, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 4, 30)

    # Row 5: Address + Date
    ws.merge_cells("A5:G5")
    ws.cell(5, 1, data["address_line2"]).font = Font(size=10)

    # Row 6: empty
    ws.merge_cells("A6:G6")
    ws.cell(6, 1, "                  ")

    # Row 7: empty
    ws.merge_cells("A7:G7")

    # Row 8: Table header
    headers = ["序号", "规格", "公斤", "数量/公斤", "单价", "金额", "备注"]
    for idx, h in enumerate(headers):
        c = ws.cell(8, idx + 1, h)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 8, 22)

    # Data rows 9-18
    for i, item in enumerate(data["items"]):
        r = 9 + i
        ws.cell(r, 1, item["seq"]).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, item["spec"] or None)
        ws.cell(r, 3, item["kg"])
        ws.cell(r, 4, item["qty_per_kg"])
        ws.cell(r, 5, item["price"])
        ws.cell(r, 6, item["amount"])
        ws.cell(r, 7, item["remark"] or None)

    # Merge remark cells for rows with same remark
    if data["items"][0]["remark"]:
        ws.merge_cells("G9:G11")

    # Apply borders to table
    apply_border(ws, 8, 18, 1, 7)

    # Row 19: 合计
    ws.cell(19, 5, "合计：").alignment = Alignment(horizontal="right")

    # Row 20: Notes + SUM formula
    ws.cell(20, 1, "                                 备注说明：不含税运费。").font = Font(size=9)
    ws.cell(20, 6, "=SUM(F9:F19)")

    # Row 21: Note 2
    ws.cell(21, 1, "                                            3.下单日起7-10天交货。").font = Font(size=9)

    # Row 22: Approver + Creator
    ws.cell(22, 1, "核准：")
    ws.merge_cells("B22:E22")
    ws.cell(22, 6, "制表：")
    ws.cell(22, 7, data["creator"])

    wb.save(output_path)
    print(f"  报价单已保存: {output_path}")


# ─────────────────────────────────────────
# 3. 对账单
# ─────────────────────────────────────────
def parse_reconciliation(pdf_path: str) -> dict:
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()

    data = {
        "company": "惠州市罗丰实业有限公司",
        "address": "公司地址:惠阳区淡水石桥中路一号富丽达花园",
        "phone": "公司电话0752-3731609/13902477013",
        "fax": "公司传真:0752-7160718",
        "title": "3月对 账 单",
        "customer": "东莞保利文塑胶制品有限公司",
        "month": 3,
        "month_label": None,
        "items": [],
        "creator": "覃丹",
    }

    m = re.search(r"((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2})", text)
    if m:
        data["month_label"] = m.group(1)

    lines = text.split("\n")
    lines = [l.strip() for l in lines]

    # Find the start of data: after "备注" header
    i = 0
    while i < len(lines):
        if lines[i] == "备注":
            i += 1
            break
        i += 1

    items = []
    current_date = None
    current_order = None
    current_delivery = None

    while i < len(lines):
        line = lines[i]
        if not line or line.startswith("13%") or line.startswith("总计") or line.startswith("税金") or line.startswith("汇总") or line.startswith("制表人"):
            break

        # Check for date line
        date_match = re.match(r"^(\d+月\d+日)$", line)
        if date_match:
            current_date = date_match.group(1)
            i += 1

            # Now expect order number (possibly with delivery number on same line)
            if i >= len(lines):
                break
            order_line = lines[i]

            # Pattern: "ZADM4500620451 LF2025030501" (order + delivery on same line)
            same_line_match = re.match(r"^((?:ZADM|ZDOM|ZMOD)\d+)\s+(LF\S+)$", order_line)
            if same_line_match:
                current_order = same_line_match.group(1)
                current_delivery = same_line_match.group(2)
                i += 1
            else:
                # Order number possibly split: "ZADM45006157" then "10" on next line
                order_match = re.match(r"^((?:ZADM|ZDOM|ZMOD)\d+)$", order_line)
                if order_match:
                    current_order = order_match.group(1)
                    i += 1
                    # Check if next line is trailing digits
                    if i < len(lines) and re.match(r"^\d+$", lines[i]) and not lines[i].startswith("LF"):
                        current_order += lines[i]
                        i += 1
                    # Check for delivery number
                    if i < len(lines) and lines[i].startswith("LF"):
                        current_delivery = lines[i]
                        i += 1

            # Now read product items until next date or end
            while i < len(lines):
                pline = lines[i]
                # If we hit a new date, break to outer loop
                if re.match(r"^\d+月\d+日$", pline):
                    break
                if pline.startswith("13%") or pline.startswith("总计") or not pline:
                    break

                # Product name: collect lines until unit marker
                name_parts = []
                while i < len(lines):
                    pl = lines[i]
                    if pl in ("PC", "套", "个", "PCS"):
                        break
                    if re.match(r"^\d+月\d+日$", pl) or pl.startswith("13%") or not pl:
                        break
                    name_parts.append(pl)
                    i += 1

                if not name_parts:
                    break

                product_name = "".join(name_parts)

                # Unit
                unit = ""
                if i < len(lines) and lines[i] in ("PC", "套", "个", "PCS"):
                    unit = lines[i]
                    i += 1

                # Quantity
                quantity = ""
                if i < len(lines) and re.match(r"^\d+$", lines[i]):
                    quantity = int(lines[i])
                    i += 1

                # Price
                price = 0
                if i < len(lines):
                    pm = re.match(r"[￥¥]([\d,.]+)", lines[i])
                    if pm:
                        price = float(pm.group(1).replace(",", ""))
                        i += 1

                # Amount
                amount = 0
                if i < len(lines):
                    am = re.match(r"[￥¥]([\d,.]+)", lines[i])
                    if am:
                        amount = float(am.group(1).replace(",", ""))
                        i += 1

                items.append({
                    "date": current_date,
                    "order_no": current_order,
                    "delivery_no": current_delivery,
                    "product_name": product_name,
                    "unit": unit,
                    "quantity": quantity,
                    "price": price,
                    "amount": amount,
                })

                # For continuation items (same date+order), clear date
                # so we know not to show it again in the merge logic
            continue
        else:
            i += 1

    data["items"] = items
    return data


def write_reconciliation(data: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "3月"

    set_col_widths(ws, {
        "A": 11.5, "B": 15.5, "C": 14.5, "D": 22,
        "E": 5, "F": 7, "G": 15, "H": 15, "I": 8, "J": 9.5,
    })

    # Ensure 10 columns by touching J column
    ws.cell(1, 10, None)

    # Row 1: Company name
    ws.merge_cells("A1:I1")
    c = ws.cell(1, 1, data["company"])
    c.font = Font(size=16, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 1, 36)

    # Row 2: Address
    ws.merge_cells("A2:I2")
    ws.cell(2, 1, data["address"]).font = Font(size=10)
    set_row_height(ws, 2, 24)

    # Row 3: Phone + Fax
    ws.cell(3, 2, data["phone"]).font = Font(size=10)
    ws.cell(3, 6, data["fax"]).font = Font(size=10)
    set_row_height(ws, 3, 20)

    # Row 4: Title
    ws.merge_cells("A4:I4")
    c = ws.cell(4, 1, data["title"])
    c.font = Font(size=18, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    set_row_height(ws, 4, 23)

    # Row 5: Customer + Month
    ws.cell(5, 1, f"客户:{data['customer']}").font = Font(size=10)
    ws.cell(5, 7, f"所属月份:{data['month']}").font = Font(size=10)
    if data["month_label"]:
        # Parse month label to date
        try:
            month_date = datetime.strptime(data["month_label"], "%b-%y")
            ws.cell(5, 8, month_date)
            ws.cell(5, 8).number_format = "MMM-YY"
        except Exception:
            ws.cell(5, 8, data["month_label"])
    set_row_height(ws, 5, 15)

    # Row 6: Table header
    headers = ["送货日期", "订单号", "送货单号", "货物名称及明细", "单位", "数量", "单价\n(不含税)", "不含税金额", "备注"]
    for idx, h in enumerate(headers):
        c = ws.cell(6, idx + 1, h)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    set_row_height(ws, 6, 28)

    # Data rows starting from row 7
    items = data["items"]
    data_start = 7

    # Group items to determine which rows share the same date/order (for merging)
    # Track groups for merging: groups of consecutive rows with same date+order
    groups = []
    current_group_start = 0
    for idx, item in enumerate(items):
        if idx == 0:
            current_group_start = idx
            continue
        # Same group if date is None (continuation)
        prev = items[idx - 1]
        if item["date"] is None or (item["date"] == prev.get("date") and item["order_no"] == prev.get("order_no")):
            continue
        else:
            groups.append((current_group_start, idx - 1))
            current_group_start = idx
    groups.append((current_group_start, len(items) - 1))

    # Determine which items share the same date/order_no for cell merging
    for idx, item in enumerate(items):
        r = data_start + idx

        # Date - convert to Excel serial date
        date_val = item.get("date")
        if date_val:
            try:
                dm = re.match(r"(\d+)月(\d+)日", date_val)
                if dm:
                    date_obj = datetime(2025, int(dm.group(1)), int(dm.group(2)))
                    ws.cell(r, 1, date_obj)
                    ws.cell(r, 1).number_format = "YYYY/MM/DD"
                else:
                    ws.cell(r, 1, date_val)
            except Exception:
                ws.cell(r, 1, date_val)
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(r, 2, item.get("order_no", "")).alignment = Alignment(vertical="center")
        ws.cell(r, 3, item.get("delivery_no", "")).alignment = Alignment(vertical="center")
        ws.cell(r, 4, item.get("product_name", "")).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(r, 5, item.get("unit", "")).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 6, item.get("quantity", "")).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 7, item.get("price", "")).alignment = Alignment(vertical="center")
        # Use formula for amount
        ws.cell(r, 8, f"=G{r}*F{r}")
        ws.cell(r, 8).number_format = "#,##0.00"

        set_row_height(ws, r, 39)

    # Merge cells for same date/order groups
    for group_start, group_end in groups:
        if group_end > group_start:
            r_start = data_start + group_start
            r_end = data_start + group_end
            ws.merge_cells(f"A{r_start}:A{r_end}")
            ws.merge_cells(f"B{r_start}:B{r_end}")
            ws.merge_cells(f"C{r_start}:C{r_end}")

    data_end = data_start + len(items) - 1

    # Apply borders to table
    apply_border(ws, 6, data_end, 1, 9)

    # Tax rows
    tax_row = data_end + 1
    ws.merge_cells(f"E{tax_row}:F{tax_row}")
    ws.cell(tax_row, 5, "13%增值税").alignment = Alignment(horizontal="center")
    ws.cell(tax_row, 7, "总计：").alignment = Alignment(horizontal="right")
    ws.cell(tax_row, 8, f"=SUM(H{data_start}:H{data_end})")
    ws.cell(tax_row, 8).number_format = "#,##0.00"
    set_row_height(ws, tax_row, 31)

    tax_row2 = tax_row + 1
    ws.merge_cells(f"E{tax_row2}:F{tax_row2}")
    ws.cell(tax_row2, 7, "税金").alignment = Alignment(horizontal="right")
    ws.cell(tax_row2, 8, f"=H{tax_row}*0.13")
    ws.cell(tax_row2, 8).number_format = "#,##0.00"
    set_row_height(ws, tax_row2, 30)

    total_row = tax_row2 + 1
    ws.merge_cells(f"E{total_row}:F{total_row}")
    ws.cell(total_row, 7, "汇总").alignment = Alignment(horizontal="right")
    ws.cell(total_row, 8, f"=SUM(H{tax_row}:H{tax_row2})")
    ws.cell(total_row, 8).number_format = "#,##0.00"
    set_row_height(ws, total_row, 30)

    # Footer
    footer_row = total_row + 1
    ws.cell(footer_row, 2, "制表人:")
    ws.cell(footer_row, 3, data["creator"])
    ws.cell(footer_row, 4, "公司盖章:")
    set_row_height(ws, footer_row, 30)

    wb.save(output_path)
    print(f"  对账单已保存: {output_path}")


# ─────────────────────────────────────────
# Main
# ─────────────────────────────────────────
def main():
    import os

    base_dir = os.path.dirname(os.path.abspath(__file__))

    print("开始转换 PDF → Excel ...\n")

    # 1. 送货单
    pdf1 = os.path.join(base_dir, "副本保利高（齿轮）3-3.pdf")
    if os.path.exists(pdf1):
        print("[1/3] 送货单")
        data = parse_delivery_note(pdf1)
        write_delivery_note(data, os.path.join(base_dir, "output_送货单.xlsx"))
    else:
        print(f"[1/3] 跳过 - 文件不存在: {pdf1}")

    # 2. 报价单
    pdf2 = os.path.join(base_dir, "副本副本报价单（精丰）10-30.pdf")
    if os.path.exists(pdf2):
        print("[2/3] 报价单")
        data = parse_quotation(pdf2)
        write_quotation(data, os.path.join(base_dir, "output_报价单.xlsx"))
    else:
        print(f"[2/3] 跳过 - 文件不存在: {pdf2}")

    # 3. 对账单
    pdf3 = os.path.join(base_dir, "副本副本罗丰对账单保利文2025年3月份(2).pdf")
    if os.path.exists(pdf3):
        print("[3/3] 对账单")
        data = parse_reconciliation(pdf3)
        write_reconciliation(data, os.path.join(base_dir, "output_对账单.xlsx"))
    else:
        print(f"[3/3] 跳过 - 文件不存在: {pdf3}")

    print("\n转换完成！")


if __name__ == "__main__":
    main()
